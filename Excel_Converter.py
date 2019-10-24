import sys
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import QApplication, QTextEdit, QWidget, QPushButton, QLabel, QLineEdit, QScrollArea, QComboBox, QMainWindow, \
    QAction, QVBoxLayout, QProgressBar,QLCDNumber,QMessageBox, QFileDialog
from PyQt5.QtGui import QIcon, QFont, QPainter, QPixmap, QIcon
from PyQt5.QtCore import pyqtSlot, Qt,QTime, QTimer, QDateTime, QDate


from datetime import datetime
import pyexcel as p
from fnmatch import fnmatch
import re
import os
import win32com.client as win32
from win32com.client import constants
from shutil import copyfile
import pyexcel as px
import pyexcel_io
import pyexcel_xls
import pyexcel_xlsx
import xlrd
from openpyxl.workbook import Workbook as openpyxlWorkbook
import time
from fnmatch import fnmatch
from PyQt5.QtCore import QThread, pyqtSignal
#file = str(QFileDialog.getExistingDirectory(self, "Select Directory"))
#########################################################################################################################
dir_name = None
class Threaded(QObject):
    global dir_name
    result=pyqtSignal(list)

    def __init__(self, parent=None, **kwargs):
        super().__init__(parent, **kwargs)

    @pyqtSlot()
    def start(self):
        print("Thread started")

    @pyqtSlot()
    def calculatePrime(self):

        #print('processing')
        #class_E =Example()
        dir_name = Threaded.dir_name
        #class_E.processDialog(dir_name)

        #self.dir_name = dir_name
        #print('0000000000000')
        #print(self.dir_name)
        if (dir_name != ""):

            ###############################################################################
            root1 = 'Reports'
            arch = 'New_archive'
            pattern = "*.xls"
            ###############################################################################
            list_total_files = []
            transferred_files_path = []
            transferred_files_name = []
            files_not_transferred = []
            files_not_copied = []
            ###############################################################################

            dir_name_new = dir_name.replace('/', '\\')

            ###############################################################################

            ###############################################################################
            counter = 1
            for path, dirs, files in os.walk(dir_name_new):
                #print(path)
                for name in files:
                    if fnmatch(name, pattern):
                        #print(name)
                        r = str(os.path.join(path, name))
                        # r= r.replace('\\','/')

                        list_total_files.append(r)
                        dest = r + "x"
                        try:
                            #p.save_book_as(file_name=r, dest_file_name=dest)
                            #transferred_files_path.append(r)
                            #transferred_files_name.append(name)
                            xlsBook = xlrd.open_workbook(r)
                            workbook = openpyxlWorkbook()

                            for i in range(0, xlsBook.nsheets):
                                xlsSheet = xlsBook.sheet_by_index(i)
                                sheet = workbook.active if i == 0 else workbook.create_sheet()
                                sheet.title = xlsSheet.name

                                for row in range(0, xlsSheet.nrows):
                                    for col in range(0, xlsSheet.ncols):
                                        sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)

                            workbook.save(dest)
                            transferred_files_path.append(r)
                            transferred_files_name.append(name)
                        except:
                            files_not_transferred.append(r)
                            #print(r + ':  Not transferred')

            if (len(transferred_files_path)) or (len(files_not_transferred)) > 0:
                arch_name = "Archive_Old_Excel_Files__"
                # datetime object containing current date and time
                now = datetime.now()
                # dd/mm/YY H:M:S
                dt_string = now.strftime("%d_%m_%Y__%H_%M_%S")
                arch_name = arch_name + dt_string
                arch_path = dir_name_new + '\\' + arch_name
                # Create target Directory if don't exist

                try:
                    os.makedirs(arch_path)
                    print("Directory ", arch_path, " Created ")
                except FileExistsError:
                    print("Directory ", arch_path, " already exists")
                # self.textEdit.setText(dir_name)
                counter = 1
                for i, k in enumerate(transferred_files_path):
                    fname = transferred_files_name[i]
                    n = arch_path + '\\_' + str(counter) + '___' + fname
                    try:
                        copyfile(k, n)
                    except:
                        files_not_copied.append(r)
                        #print(r+  ':  not copied')

                    try:
                        os.remove(k)
                    except:
                        print('file cannot be removed')
                    counter = counter + 1

                ###############################################################################
                name_file = dir_name + '\\' + arch_name + '\\' + 'Converted_Files' + '.txt'
                name_file_2 = dir_name + '\\' + arch_name + '\\' + 'Not_Converted_Files.txt'
                f1 = open(name_file, "w+")
                f2 = open(name_file_2, "w+")
                f2.write('=' * 250)
                f2.write('\n')
                f2.write('                  File Name                         \n')
                f2.write('=' * 250)
                f2.write('\n\n')
                for i in files_not_transferred:
                    f2.write(i)
                    f2.write("\n")
                ###############################################################################
                f1.write('=' * 250)
                f1.write('\n')
                f1.write('               File Name                      ||||                Complete Path   \n')
                f1.write('=' * 250)
                f1.write('\n\n')
                counter = 1
                for i, k in enumerate(transferred_files_path):
                    f1.write(str(counter) + '___' + transferred_files_name[i])
                    f1.write('  ||||  ')
                    f1.write(k)
                    f1.write("\n")
                    counter = counter + 1

                Text_Displayed = '\n\n' + ('=' * 70)
                Text_Displayed = Text_Displayed + '\n        Information\n'
                Text_Displayed = Text_Displayed + ('=' * 70)
                Text_Displayed = Text_Displayed + '\n\nTotal Number of .xls Files found: ' + str(len(list_total_files))
                Text_Displayed = Text_Displayed + '\n\nNumber of Files converted: ' + str(len(transferred_files_path))
                Text_Displayed = Text_Displayed + '\n\nNumber of Files cannot be converted: ' + str(
                    len(files_not_transferred))
                Text_Displayed_2 = '\n\n\n\nPath of the Archive Directory files :   \n' + ('-' * 100) + '\n'
                Text_Displayed_3 = arch_path

                n = 100
                length_arch_path = len(arch_path)
                if (length_arch_path > 110):
                    text_array = [Text_Displayed_3[i:i + n] for i in range(0, len(Text_Displayed_3), n)]
                    Text_Displayed_3 = ""
                    for each in text_array:
                        Text_Displayed_3 = Text_Displayed_3 + '\n' + each
                # self.label_2.setText(Text_Displayed)
                # self.label_3.setText(Text_Displayed_2)
                # self.label_4.setText(Text_Displayed_3)
                list_displayed = []
                list_displayed.append(Text_Displayed)
                list_displayed.append(Text_Displayed_2)
                list_displayed.append(Text_Displayed_3)

            else:
                text5 = '\n\nNote: No Excel Files found with .xls extension'
                list_displayed = []
                list_displayed.append(text5)
        self.result.emit(list_displayed)

##########################################################################################################################
class Example(QMainWindow):
    global dir_name
    requestPrime = pyqtSignal()
    def __init__(self):
        super().__init__()

        self.initUI()
        self._thread = QThread(self)
        self._threaded = Threaded(result=self.endDialog)
        self.requestPrime.connect(self._threaded.calculatePrime)
        self._thread.started.connect(self._threaded.start)
        self._threaded.moveToThread(self._thread)
        #qApp.aboutToQuit.connect(self._thread.quit)

    def initUI(self):


        self.progressbar = QProgressBar(self)
        # self.progressbar.setOrientation(Qt.Vertical)
        self.progressbar.setMaximum(100)
        self.progressbar.setStyleSheet("QProgressBar {border: 2px solid grey;border-radius:8px;padding:1px}"
                                       "QProgressBar::chunk {background:green}")
        self.progressbar.setGeometry(100, 800, 411, 30)
        self.progressbar.setProperty("value", 0)
        self.progressbar.setTextVisible(False)
        self.progressbar.hide()

        self.label_p = QLabel(self)
        self.label_p.setFont(QFont("Sans Serif", 14, QFont.Bold))
        #self.label_1.setStyleSheet("QLabel{color:darkblue;background:darkgray;border-radius: 5px; border: 1px ; }")
        self.label_p.setText('         NOTE: Kindly Wait until the processing is being completed')
        self.label_p.setGeometry(50, 700, 700, 100)
        self.label_p.hide()



        #self.centralwidget = QWidget(self)
        self.label_1 = QLabel(self)
        self.label_1.setFont(QFont("Sans Serif", 24, QFont.Bold))
        self.label_1.setStyleSheet(
            "QLabel{color:darkblue;background:darkgray;border-radius: 5px; border: 1px ; }")
        self.label_1.setText('                          Infineon Excel File Converter')
        self.label_1.setGeometry(10, 45, 900, 80)

        self.label_2 = QLabel(self)
        self.label_2.setFont(QFont("Sans Serif", 16))
        self.label_2.setGeometry(10, 130, 900, 300)
        #self.label_2.setAlignment(Qt.AlignVTop)
        self.label_2.setAlignment(Qt.AlignTop)
        self.label_2.setTextInteractionFlags(Qt.TextSelectableByMouse)
        #self.label_2.setStyleSheet("QLabel{color:black;background:darkgray; }")
        self.label_2.setMinimumSize(self.sizeHint())


        self.label_3 = QLabel(self)
        self.label_3.setFont(QFont("Sans Serif", 14, QFont.Bold))
        self.label_3.setGeometry(10, 440, 900, 130)
        #self.label_3.setAlignment(Qt.AlignVTop)
        self.label_3.setAlignment(Qt.AlignTop)
        self.label_3.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.label_3.setMargin(0)

        #self.label_3.setStyleSheet("QLabel{color:darkblue;background:darkgray; padding:0 0 0 0px}")
        self.label_3.setMinimumSize(self.sizeHint())

        self.label_4 = QLabel(self)
        self.label_4.setFont(QFont("Sans Serif", 14))
        self.label_4.setGeometry(10, 575, 900, 200)
        self.label_4.setWordWrap(True)
        #self.setCentralWidget(self.textEdit)
        self.label_4.setAlignment(Qt.AlignTop)
        self.label_4.setTextInteractionFlags(Qt.TextSelectableByMouse)
        #self.label_4.setStyleSheet("QLabel{color:darkblue;background:darkgray; }")
        #self.label_4.setAlignment(Qt.A)

        self.statusBar()
        self.setWindowIcon(QIcon('go.png'))
        openFile = QAction(QIcon('open.png'), 'Open', self)
        openFile.setShortcut('Ctrl+O')
        openFile.setStatusTip('Select a Directory')
        openFile.triggered.connect(self.showDialog)

        menubar = self.menuBar()
        fileMenu = menubar.addMenu('&File')
        fileMenu.addAction(openFile)

        helpMenu = menubar.addMenu('&Help')
        help = QAction('About the Tool', self)
        helpMenu.addAction(help)
        helpMenu.triggered.connect(self.helpDialog)
        self.setFixedSize(920, 900)
        self.setWindowTitle('Infineon\'s Excel Converter')


        self.setStyleSheet("""
                           QMainWindow{
                                Background: silver;
                                color:white;
                                font:18px bold;
                                font-weight:bold;
                                border-radius: 50px ;
                                height: 11;
                           }
                           QMenuBar {
                                 background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                    stop:0 lightgray, stop:1 darkgray);
                                                    font-size:16px;
                                                    font-weight: bold;

                           }

                           QMenuBar::item {
                               spacing: 300px;           
                               padding: 8px 40px;
                               background-color: rgb(0,0,139);
                               color: rgb(255,255,255);  
                               border-radius: 5px;
                               border : 2px  solid black;
                               font-size:19px;


                           }

                           QMenuBar::item::selected {
                               background-color: rgb(30,30,30);

                           }

                           QMenu {
                               background-color: rgb(49,49,49);
                               color: rgb(0,255,255);



                           }
                           QMenu::item {
                               spacing: 30px;           
                               padding: 7px 40px;
                               font-size:16px;
                               background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                                    stop:0 lightgray, stop:1 darkgray);
                               color: rgb(0,0,0);  
                               border-radius: 2px;
                               border: 2px solid white; 

                           }

                           QMenu::item::selected {
                               background-color: rgb(49,49,49);
                               color: rgb(255,255,255);
                           }
                           QMenu::item::pressed {
                               background-color: rgb(194,105,30);
                               color: rgb(255,255,255);
                           }
                       """)
        self.show()

    def helpDialog(self):
        text = '\nThe tool automatically convert all .xls files into .xlsx files within a specific Directory or Folder.'
        text = text + '\nThe tool also create an Archive folder within the same directory having all the original .xls files.'
        text= text + '\n\n' + '='*50
        text= text + '\n  Steps\n'
        text = text + '='*50 +'\nStep1: Click on File Menu\nStep2: Click on Open\nStep3: Select a Directory containing .xls files'
        self.label_2.setText(text)
        self.label_3.setText('')
        self.label_4.setText('')

    @pyqtSlot()
    def kkDialog(self, dir_name):
        print('function called')
        print(dir_name)
    @pyqtSlot()
    def processDialog(self,dir_name):
        self.dir_name= dir_name
        #print('0000000000000')
        #print(self.dir_name)
        if(dir_name!=""):

            ###############################################################################
            root1 = 'Reports'
            arch = 'New_archive'
            pattern = "*.xls"
            ###############################################################################
            list_total_files= []
            transferred_files_path = []
            transferred_files_name = []
            files_not_transferred = []
            files_not_copied = []
            ###############################################################################


            dir_name_new= dir_name.replace('/','\\')
            #self.textEdit_1.setText(dir_name_new)
            ###############################################################################



            ###############################################################################
            counter = 1
            for path, dirs, files in os.walk(dir_name_new):
                #print(path)
                for name in files:
                    if fnmatch(name, pattern):
                        print(name)
                        r = str(os.path.join(path, name))
                        #r= r.replace('\\','/')

                        list_total_files.append(r)
                        dest = r + "x"
                        try:
                            p.save_book_as(file_name=r, dest_file_name=dest)
                            transferred_files_path.append(r)
                            transferred_files_name.append(name)
                        except:
                            files_not_transferred.append(r)





            if (len(transferred_files_path)) or (len(files_not_transferred)) > 0 :
                arch_name = "Archive_Old_Excel_Files__"
                # datetime object containing current date and time
                now = datetime.now()
                # dd/mm/YY H:M:S
                dt_string = now.strftime("%d_%m_%Y__%H_%M_%S")
                arch_name = arch_name + dt_string
                arch_path = dir_name_new + '\\' + arch_name
                # Create target Directory if don't exist

                try:
                    os.makedirs(arch_path)
                    print("Directory ", arch_path, " Created ")
                except FileExistsError:
                    print("Directory ", arch_path, " already exists")
                #self.textEdit.setText(dir_name)
                counter = 1
                for i,k in enumerate(transferred_files_path):
                    fname = transferred_files_name[i]
                    n = arch_path + '\\_' + str(counter) + '___' + fname
                    try:
                        copyfile(k, n)
                    except:
                        files_not_copied.append(r)

                    try:
                        os.remove(k)
                    except:
                        print('file cannot be removed')
                    counter = counter + 1


                ###############################################################################
                name_file = dir_name + '\\' + arch_name + '\\' + 'Converted_Files' + '.txt'
                name_file_2 = dir_name + '\\' + arch_name + '\\' + 'Not_Converted_Files.txt'
                f1 = open(name_file, "w+")
                f2 = open(name_file_2, "w+")
                f2.write('=' * 250)
                f2.write('\n')
                f2.write('                  File Name                         \n')
                f2.write('=' * 250)
                f2.write('\n\n')
                for i in files_not_transferred:
                    f2.write(i)
                    f2.write("\n")
                ###############################################################################
                f1.write('='*250)
                f1.write('\n')
                f1.write('               File Name                      ||||                Complete Path   \n')
                f1.write('=' * 250)
                f1.write('\n\n')
                counter= 1
                for i,k in enumerate(transferred_files_path):
                    f1.write(str(counter)+'___'+ transferred_files_name[i])
                    f1.write('  ||||  ')
                    f1.write(k)
                    f1.write("\n")
                    counter= counter + 1

                Text_Displayed =  '=' * 70
                Text_Displayed= Text_Displayed + '\n        Information\n'
                Text_Displayed = Text_Displayed + ('=' * 70)
                Text_Displayed = Text_Displayed + '\nTotal Number of .xls Files found: '+ str(len(list_total_files))
                Text_Displayed = Text_Displayed + '\nNumber of Files converted: ' + str(len(transferred_files_path))
                Text_Displayed = Text_Displayed + '\nNumber of Files cannot be converted: ' + str(len(files_not_transferred))
                #Text_Displayed_2 = 'Path of the Archive Directory files :   \n'+ ('-' * 100)+'\n'
                Text_Displayed_2 = 'Path of the Archive Directory files :'
                Text_Displayed_3 =  arch_path

                n = 100
                length_arch_path = len(arch_path)
                if(length_arch_path>110):
                    text_array= [Text_Displayed_3[i:i + n] for i in range(0, len(Text_Displayed_3), n)]
                    Text_Displayed_3 = ""
                    for each in text_array:
                        Text_Displayed_3= Text_Displayed_3+ '\n'+ each
                #self.label_2.setText(Text_Displayed)
                #self.label_3.setText(Text_Displayed_2)
                #self.label_4.setText(Text_Displayed_3)
                list_displayed=[]
                list_displayed.append(Text_Displayed)
                list_displayed.append(Text_Displayed_2)
                list_displayed.append(Text_Displayed_3)

            else:
                text5= '\n\nNote: No Excel Files found with .xls extension'
                list_displayed = []
                list_displayed.append(text5)
                #self.label_2.setText('\n\nNote: No Excel Files found with .xls extension')


    @pyqtSlot()
    def showDialog(self):

        dir_name = str(QFileDialog.getExistingDirectory(self, 'Select Directory'))
        self._thread.start()
        self.label_2.setText('')
        self.progressbar.setRange(0, 0)
        self.progressbar.show()
        self.label_p.show()
        print(dir_name)
        Threaded.dir_name = dir_name
        print('continue')
        self.requestPrime.emit()
        #self.git_thread.start()
        #self.processDialog()
    @pyqtSlot(list)
    def endDialog(self,list):
        if len(list)==1:
            self.label_2.setText('\n\nNote: No Excel Files found with .xls extension')
        else:
            self.label_2.setText(list[0])
            self.label_3.setText(list[1])
            self.label_4.setText(list[2])
        self.progressbar.setRange(0, 100)
        self.progressbar.hide()
        self.label_p.hide()
        print('Completed task')
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Example()
    sys.exit(app.exec_())